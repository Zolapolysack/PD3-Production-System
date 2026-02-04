# -*- coding: utf-8 -*-
"""
สคริปต์สำหรับสร้าง Excel Template ที่ตรงกับ TEST_TEMPLATE_FINAL.xlsx 100%
วิเคราะห์และสร้างจากไฟล์ที่ผู้ใช้แก้ไขให้สมบูรณ์แล้ว
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

def create_perfect_template():
    """สร้าง template ที่ตรงกับ TEST_TEMPLATE_FINAL.xlsx ทุกรายละเอียด"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # ตั้งชื่อ sheet
    date_str = datetime.now().strftime('%d-%m-%Y')
    ws.title = f'รายงาน {date_str}'
    
    # ===== COLUMN WIDTHS =====
    ws.column_dimensions['A'].width = 15.8984375
    ws.column_dimensions['B'].width = 18.0
    ws.column_dimensions['C'].width = 13.0
    ws.column_dimensions['D'].width = 13.0
    ws.column_dimensions['E'].width = 12.09765625
    ws.column_dimensions['F'].width = 13.0
    ws.column_dimensions['G'].width = 19.0
    ws.column_dimensions['H'].width = 13.0
    ws.column_dimensions['I'].width = 13.0
    ws.column_dimensions['J'].width = 13.0
    ws.column_dimensions['K'].width = 13.0
    ws.column_dimensions['L'].width = 13.0
    
    # ===== ROW HEIGHTS =====
    for row_idx in range(7, 39):
        ws.row_dimensions[row_idx].height = 19.8
    
    # ===== FONT DEFINITIONS =====
    font_arial_11_bold = Font(name='Arial', size=11, bold=True)
    font_arial_11 = Font(name='Arial', size=11, bold=False)
    font_tahoma_11 = Font(name='Tahoma', size=11, bold=False)
    font_angsana_14 = Font(name='Angsana New', size=14, bold=False)
    
    # ===== MERGED CELLS =====
    merge_ranges = [
        'B1:C1', 'D1:E1', 'F1:J1',
        'B2:C2', 'D2:E2', 'F2:L2',
        'D3:E3', 'F3:L3',
        'D4:E4', 'F4:L4',
        'C5:F5', 'H5:L5',
        'E6:F6', 'K6:L6',
        # PT rows
        'A7:A10', 'G7:G10',
        'A11:A14', 'G11:G14',
        'A15:A18', 'G15:G18',
        'A19:A22', 'G19:G22',
        'A23:A26', 'G23:G26',
        'A27:A30', 'G27:G30',
        'A31:A34', 'G31:G34',
        # หมายเหตุ columns
        'E7:F7', 'E8:F8', 'E9:F9', 'E10:F10',
        'E11:F11', 'E12:F12', 'E13:F13', 'E14:F14',
        'E15:F15', 'E16:F16', 'E17:F17', 'E18:F18',
        'E19:F19', 'E20:F20', 'E21:F21', 'E22:F22',
        'E23:F23', 'E24:F24', 'E25:F25', 'E26:F26',
        'E27:F27', 'E28:F28', 'E29:F29', 'E30:F30',
        'E31:F31', 'E32:F32', 'E33:F33', 'E34:F34',
        'K7:L7', 'K8:L8', 'K9:L9', 'K10:L10',
        'K11:L11', 'K12:L12', 'K13:L13', 'K14:L14',
        'K15:L15', 'K16:L16', 'K17:L17', 'K18:L18',
        'K19:L19', 'K20:L20', 'K21:L21', 'K22:L22',
        'K23:L23', 'K24:L24', 'K25:L25', 'K26:L26',
        'K27:L27', 'K28:L28', 'K29:L29', 'K30:L30',
        'K31:L31', 'K32:L32', 'K33:L33', 'K34:L34',
        'J35:L35',
        'I37:J37'
    ]
    
    for merge_range in merge_ranges:
        ws.merge_cells(merge_range)
    
    # ===== ROW 1 =====
    ws['A1'].font = font_arial_11_bold
    ws['A1'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), 
                             left=Side(style='thin'), right=Side(style='thin'))
    
    ws['B1'] = 'ยอดพิมพ์(ใบ)'
    ws['B1'].font = font_arial_11_bold
    ws['B1'].alignment = Alignment(horizontal='center', vertical='top')
    ws['B1'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['C1'].font = font_tahoma_11
    ws['C1'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['D1'] = 'จำนวนพนักงานขาด'
    ws['D1'].font = font_arial_11_bold
    ws['D1'].alignment = Alignment(horizontal='center')
    ws['D1'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['E1'].font = font_tahoma_11
    ws['E1'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['F1'] = 'หมายเหตุ/ปัญหาที่พบ'
    ws['F1'].font = font_arial_11_bold
    ws['F1'].alignment = Alignment(horizontal='center')
    ws['F1'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in ['G', 'H', 'I', 'J', 'K']:
        ws[f'{col}1'].font = font_tahoma_11
        ws[f'{col}1'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['L1'].font = font_tahoma_11
    ws['L1'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    # ===== ROW 2 =====
    ws['A2'] = 'ยอดยกมา'
    ws['A2'].font = font_arial_11_bold
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), 
                             left=Side(style='thin'), right=Side(style='thin'))
    
    ws['B2'].font = font_arial_11
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['B2'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['C2'].font = font_tahoma_11
    ws['C2'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['D2'].font = font_arial_11_bold
    ws['D2'].alignment = Alignment(horizontal='center')
    ws['D2'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['E2'].font = font_tahoma_11
    ws['E2'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['F2'].font = font_arial_11_bold
    ws['F2'].alignment = Alignment(horizontal='center')
    ws['F2'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    for col in ['G', 'H', 'I', 'J', 'K']:
        ws[f'{col}2'].font = font_tahoma_11
        ws[f'{col}2'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['L2'].font = font_tahoma_11
    ws['L2'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    # ===== ROW 3 =====
    ws['A3'] = 'ยอดผลิตประจำวัน'
    ws['A3'].font = font_arial_11_bold
    ws['A3'].alignment = Alignment(horizontal='center')
    ws['A3'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), 
                             left=Side(style='thin'), right=Side(style='thin'))
    
    ws['B3'] = '=D36+J36'
    ws['B3'].font = font_arial_11_bold
    ws['B3'].alignment = Alignment(horizontal='center')
    ws['B3'].border = Border(bottom=Side(style='thin'))
    
    ws['C3'].font = font_arial_11_bold
    ws['C3'].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['D3'].font = font_arial_11_bold
    ws['D3'].alignment = Alignment(horizontal='center')
    ws['D3'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['E3'].font = font_tahoma_11
    ws['E3'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['F3'].font = font_arial_11_bold
    ws['F3'].alignment = Alignment(horizontal='center')
    ws['F3'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    for col in ['G', 'H', 'I', 'J', 'K']:
        ws[f'{col}3'].font = font_tahoma_11
        ws[f'{col}3'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['L3'].font = font_tahoma_11
    ws['L3'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    # ===== ROW 4 =====
    ws['A4'] = 'รวม'
    ws['A4'].font = font_arial_11_bold
    ws['A4'].alignment = Alignment(horizontal='center')
    ws['A4'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), 
                             left=Side(style='thin'), right=Side(style='thin'))
    
    ws['C4'].font = font_arial_11_bold
    ws['C4'].border = Border(right=Side(style='thin'))
    
    ws['D4'].font = font_arial_11_bold
    ws['D4'].alignment = Alignment(horizontal='center')
    ws['D4'].border = Border(top=Side(style='thin'), right=Side(style='thin'))
    
    ws['E4'].font = font_tahoma_11
    ws['E4'].border = Border(top=Side(style='thin'), right=Side(style='thin'))
    
    ws['F4'].font = font_arial_11_bold
    ws['F4'].alignment = Alignment(horizontal='center')
    ws['F4'].border = Border(top=Side(style='thin'), right=Side(style='thin'))
    
    for col in ['G', 'H', 'I', 'J', 'K']:
        ws[f'{col}4'].font = font_tahoma_11
        ws[f'{col}4'].border = Border(top=Side(style='thin'))
    
    ws['L4'].font = font_tahoma_11
    ws['L4'].border = Border(top=Side(style='thin'), right=Side(style='thin'))
    
    # ===== ROW 5 =====
    ws['A5'].font = font_arial_11_bold
    ws['A5'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
    
    ws['B5'] = 'ยอดผลิตประจำวันที่'
    ws['B5'].font = font_arial_11_bold
    ws['B5'].alignment = Alignment(horizontal='center')
    ws['B5'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
    
    ws['C5'].font = font_arial_11_bold
    ws['C5'].alignment = Alignment(horizontal='center')
    ws['C5'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    for col in ['D', 'E']:
        ws[f'{col}5'].font = font_tahoma_11
        ws[f'{col}5'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['F5'].font = font_tahoma_11
    ws['F5'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['G5'] = 'ยอดผลิตประจำวันที่'
    ws['G5'].font = font_arial_11_bold
    ws['G5'].alignment = Alignment(horizontal='left')
    ws['G5'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
    
    ws['H5'].font = font_arial_11_bold
    ws['H5'].alignment = Alignment(horizontal='center')
    ws['H5'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    for col in ['I', 'J', 'K']:
        ws[f'{col}5'].font = font_tahoma_11
        ws[f'{col}5'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['L5'].font = font_tahoma_11
    ws['L5'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    # ===== ROW 6 (Headers) =====
    ws['A6'] = 'PT'
    ws['A6'].font = font_arial_11_bold
    ws['A6'].alignment = Alignment(horizontal='center')
    ws['A6'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), 
                             left=Side(style='thin'), right=Side(style='thin'))
    
    ws['B6'] = 'ตรา'
    ws['B6'].font = font_arial_11_bold
    ws['B6'].alignment = Alignment(horizontal='center')
    ws['B6'].border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
    
    ws['C6'] = 'เวลา'
    ws['C6'].font = font_arial_11_bold
    ws['C6'].alignment = Alignment(horizontal='center')
    ws['C6'].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['D6'] = 'จำนวน (ใบ)'
    ws['D6'].font = font_arial_11_bold
    ws['D6'].alignment = Alignment(horizontal='center')
    ws['D6'].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['E6'] = 'เลขที่ S/O'
    ws['E6'].font = font_arial_11_bold
    ws['E6'].alignment = Alignment(horizontal='center')
    ws['E6'].border = Border(bottom=Side(style='thin'))
    
    ws['F6'].font = font_tahoma_11
    ws['F6'].border = Border(bottom=Side(style='thin'))
    
    ws['G6'] = 'PT'
    ws['G6'].font = font_arial_11_bold
    ws['G6'].alignment = Alignment(horizontal='center')
    ws['G6'].border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
    
    ws['H6'] = 'ตรา'
    ws['H6'].font = font_arial_11_bold
    ws['H6'].alignment = Alignment(horizontal='center')
    ws['H6'].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['I6'] = 'เวลา'
    ws['I6'].font = font_arial_11_bold
    ws['I6'].alignment = Alignment(horizontal='center')
    ws['I6'].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['J6'] = 'จำนวน (ใบ)'
    ws['J6'].font = font_arial_11_bold
    ws['J6'].alignment = Alignment(horizontal='center')
    ws['J6'].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['K6'] = 'เลขที่ S/O'
    ws['K6'].font = font_arial_11_bold
    ws['K6'].alignment = Alignment(horizontal='center')
    ws['K6'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    ws['L6'].font = font_tahoma_11
    ws['L6'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    # ===== DATA ROWS (7-34) =====
    pt_configs = [
        (1, 7),   # PT 1: rows 7-10
        (2, 11),  # PT 2: rows 11-14
        (3, 15),  # PT 3: rows 15-18
        (4, 19),  # PT 4: rows 19-22
        (8, 23),  # PT 8: rows 23-26
        (9, 27),  # PT 9: rows 27-30
        (10, 31)  # PT 10: rows 31-34
    ]
    
    for pt_num, start_row in pt_configs:
        # PT number in column A (merged)
        ws[f'A{start_row}'] = pt_num
        ws[f'A{start_row}'].font = font_angsana_14
        ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{start_row}'].border = Border(
            top=Side(style='thin'), bottom=Side(style='thin'),
            left=Side(style='thin'), right=Side(style='thin')
        )
        
        # PT number in column G (merged)
        ws[f'G{start_row}'] = pt_num
        ws[f'G{start_row}'].font = font_angsana_14
        ws[f'G{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'G{start_row}'].border = Border(
            top=Side(style='thin'), bottom=Side(style='thin'),
            left=Side(style='thin'), right=Side(style='thin')
        )
        
        # Data cells for 4 rows
        for offset in range(4):
            row = start_row + offset
            is_first_row = (offset == 0)
            
            # Column A (merged, already set above for first row)
            if not is_first_row:
                ws[f'A{row}'].font = font_tahoma_11
                ws[f'A{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'))
            
            # Column B (ตรา)
            ws[f'B{row}'].font = font_angsana_14
            ws[f'B{row}'].alignment = Alignment(horizontal='left')
            ws[f'B{row}'].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
            
            # Column C (เวลา)
            ws[f'C{row}'].font = font_angsana_14
            if is_first_row:
                ws[f'C{row}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), 
                                              left=Side(style='thin'), right=Side(style='thin'))
            else:
                ws[f'C{row}'].border = Border(bottom=Side(style='thin'), 
                                              left=Side(style='thin'), right=Side(style='thin'))
            
            # Column D (จำนวน)
            ws[f'D{row}'].font = font_angsana_14
            if is_first_row:
                ws[f'D{row}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), 
                                              left=Side(style='thin'), right=Side(style='thin'))
            else:
                ws[f'D{row}'].border = Border(bottom=Side(style='thin'), 
                                              left=Side(style='thin'), right=Side(style='thin'))
            
            # Columns E:F (หมายเหตุ - merged)
            ws[f'E{row}'].font = font_angsana_14
            ws[f'E{row}'].alignment = Alignment(horizontal='center')
            if is_first_row:
                ws[f'E{row}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
            else:
                ws[f'E{row}'].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
            
            ws[f'F{row}'].font = font_tahoma_11
            if is_first_row:
                ws[f'F{row}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
            else:
                ws[f'F{row}'].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
            
            # Column G (merged, already set above for first row)
            if not is_first_row:
                ws[f'G{row}'].font = font_tahoma_11
                ws[f'G{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'))
            
            # Column H (ตรา)
            ws[f'H{row}'].font = font_angsana_14
            ws[f'H{row}'].alignment = Alignment(horizontal='left')
            if row == 34:  # Last row special case
                ws[f'H{row}'].border = Border(right=Side(style='thin'))
            else:
                ws[f'H{row}'].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
            
            # Column I (เวลา)
            ws[f'I{row}'].font = font_angsana_14
            if is_first_row:
                ws[f'I{row}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), 
                                              left=Side(style='thin'), right=Side(style='thin'))
            elif row == 34:
                ws[f'I{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'))
            else:
                ws[f'I{row}'].border = Border(bottom=Side(style='thin'), 
                                              left=Side(style='thin'), right=Side(style='thin'))
            
            # Column J (จำนวน)
            ws[f'J{row}'].font = font_angsana_14
            if is_first_row:
                ws[f'J{row}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), 
                                              left=Side(style='thin'), right=Side(style='thin'))
            elif row == 34:
                ws[f'J{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'))
            else:
                ws[f'J{row}'].border = Border(bottom=Side(style='thin'), 
                                              left=Side(style='thin'), right=Side(style='thin'))
            
            # Columns K:L (หมายเหตุ - merged)
            ws[f'K{row}'].font = font_angsana_14
            ws[f'K{row}'].alignment = Alignment(horizontal='center')
            if is_first_row:
                ws[f'K{row}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
            elif row == 34:
                ws[f'K{row}'].border = Border(right=Side(style='thin'))
            else:
                ws[f'K{row}'].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
            
            ws[f'L{row}'].font = font_tahoma_11
            if is_first_row:
                ws[f'L{row}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
            elif row == 34:
                ws[f'L{row}'].border = Border(right=Side(style='thin'))
            else:
                ws[f'L{row}'].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    
    # Fix specific row 10 and 34 bottom borders
    for row in [10, 14, 18, 22, 26, 30]:
        ws[f'A{row}'].border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        ws[f'G{row}'].border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
    
    # ===== ROW 35 =====
    ws['A35'] = 'รวม'
    ws['A35'].font = font_angsana_14
    ws['A35'].alignment = Alignment(horizontal='left')
    ws['A35'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
    
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws[f'{col}35'].font = font_tahoma_11
        ws[f'{col}35'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['J35'].font = font_tahoma_11
    ws['J35'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    for col in ['K']:
        ws[f'{col}35'].font = font_tahoma_11
        ws[f'{col}35'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['L35'].font = font_tahoma_11
    ws['L35'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    
    # ===== ROW 36 =====
    ws['D36'] = '=SUM(D7:D35)'
    ws['D36'].font = font_angsana_14
    ws['D36'].alignment = Alignment(horizontal='right')
    
    ws['H36'] = 'ลงชื่อ'
    ws['H36'].font = font_angsana_14
    ws['H36'].alignment = Alignment(horizontal='right')
    
    ws['J36'] = '=SUM(J7:J35)'
    ws['J36'].font = font_angsana_14
    ws['J36'].alignment = Alignment(horizontal='right')
    
    # ===== ROW 37 =====
    ws['I37'] = 'หัวหน้าแผนก PD3'
    ws['I37'].font = font_angsana_14
    ws['I37'].alignment = Alignment(horizontal='left')
    
    # ===== ROW 38 =====
    ws['I38'] = '…../…../…..'
    ws['I38'].font = font_angsana_14
    ws['I38'].alignment = Alignment(horizontal='center')
    
    # บันทึกไฟล์
    filename = 'PERFECT_TEMPLATE.xlsx'
    wb.save(filename)
    print(f'✅ สร้างไฟล์สำเร็จ: {filename}')
    print('\n📋 รายละเอียด:')
    print(f'  - Column widths: ตรงตาม TEST_TEMPLATE_FINAL.xlsx')
    print(f'  - Row heights: 19.8 สำหรับแถว 7-38')
    print(f'  - Merged cells: {len(merge_ranges)} ranges')
    print(f'  - Fonts: Arial, Tahoma, Angsana New')
    print(f'  - PT numbers: Integer (1, 2, 3, 4, 8, 9, 10)')
    print(f'  - Formulas: =D36+J36, =SUM(D7:D35), =SUM(J7:J35)')
    print(f'  - Borders: ทุก cell ตามต้นฉบับ 100%')
    
if __name__ == '__main__':
    print('=' * 80)
    print('กำลังสร้าง Perfect Template จาก TEST_TEMPLATE_FINAL.xlsx')
    print('=' * 80)
    create_perfect_template()
